import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def process_workbook(from_file, to_file):
    wb = xl.load_workbook(from_file)
    sheet = wb['Аркуш1']

    # cell = sheet.cell(1, 1)  # same as -> cell = sheet['a1']

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet,
              min_row=2,
              max_row=sheet.max_row,
              min_col=4,
              max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    wb.save(to_file)


process_workbook('../price.xlsx', '../price_corrected.xlsx')
